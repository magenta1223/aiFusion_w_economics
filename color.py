from glob import glob
import openpyxl
import pandas as pd
from string import ascii_uppercase
from openpyxl.styles import PatternFill

def color():
    pivotTables = glob("./results_pivot*.csv")

    with pd.ExcelWriter("./pivotAggregated.xlsx") as writer:

        for t in pivotTables:
            sheet_name = ' '.join(t.split("_")[2:]).replace(".csv", "")
            pd.read_csv(t).to_excel(writer, sheet_name = sheet_name, index = False, header= None)



    workbook = openpyxl.load_workbook("./pivotAggregated.xlsx")
    FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    for sheetname in workbook.sheetnames:

        sheet = workbook[sheetname]
        colRange = ascii_uppercase[sheet.max_column-1]
        rowRange = sheet.max_row


        for row in sheet["C3" : f"{colRange}{rowRange}"]:
            for cell in row:
                if "CW" in sheetname:
                    threshold = 0.5
                else:
                    threshold = 1

                if cell.value >= threshold:
                    cell.fill = FILL
                
    workbook.save("./pivotAggregated.xlsx")